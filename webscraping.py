import sys
import codecs
sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import time
import pandas as pd
import xlsxwriter
import openpyxl

service = Service(executable_path="chromedriver.exe")
driver = webdriver.Chrome(service=service)

driver.get("https://porta.fda.moph.go.th/FDA_SEARCH_ALL/MAIN/SEARCH_CENTER_MAIN.aspx")

input_element = driver.find_element(By.CLASS_NAME, "input-lg")
A = "ครีมไพล"
input_element.send_keys(A + Keys.ENTER)

checkbox_element = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.ID, "ContentPlaceHolder1_CheckBoxList1_2"))
)
checkbox_element.click()
search = driver.find_element(By.CLASS_NAME, "btn-lg")
search.click()

try :
    combobox = driver.find_element(By.NAME ,"ctl00$ContentPlaceHolder1$RadGrid1$ctl00$ctl03$ctl01$PageSizeComboBox")
    combobox.click()
    combobox.send_keys(Keys.ARROW_DOWN)
    combobox.send_keys(Keys.ARROW_DOWN)
    combobox.send_keys(Keys.ENTER)
    time.sleep(5)
except :
    time.sleep(5)
    #time.sleep(10)

# Get the total number of pages
try:
    total_pages_text = driver.find_element(By.XPATH, '/html/body/form/div[3]/div[2]/div/div/div[2]/div[2]/div/center/div/table/tfoot/tr/td/table/tbody/tr/td/div[5]/strong[2]').text
    total_pages = int(total_pages_text)
except NoSuchElementException:
    print("1 page")
    total_pages = 1

data = []
for page_num in range(1, total_pages + 1): 
    tbody = driver.find_element(By.XPATH, '/html/body/form/div[3]/div[2]/div/div/div[2]/div[2]/div/center/div/table/tbody')
    for tr in tbody.find_elements(By.XPATH, '//tr'): 
        row = [item.text for item in tr.find_elements(By.XPATH, './/td')]
        try:
            idx16 = tr.find_element(By.LINK_TEXT,value = 'ดูข้อมูล').get_attribute("href")
        except NoSuchElementException:
            print("not found link")
        for index,item in enumerate(row):
            if row[index] == "ดูข้อมูล":
                print("found at idx {} = {}".format(index,item))
                # row[index] = idx16
                driver.switch_to.new_window()
                driver.get(idx16)
                tabs_value = driver.window_handles
                print(tabs_value)
                current_value = driver.current_window_handle
                print(current_value)
                value1 = driver.find_element(By.ID,"ContentPlaceHolder1_lb_sale_channel")   #ช่องทางการขาย
                val1txt = value1.text
                value2 = driver.find_element(By.ID,"ContentPlaceHolder1_lb_indication")   #ข้อบ่งใช้
                val2txt = value2.text
                value3 = driver.find_element(By.ID,"ContentPlaceHolder1_lb_thanm")     #ชื่อบริษัท
                val3txt = value3.text
                print("value1 = {} value2 = {} value2 = {} ".format(val1txt,val2txt,val3txt))
                row[index] = val1txt
                row.append(val2txt)
                row.append(val3txt)
                print(row)
                time.sleep(2)
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
        #if look last colomn
        #last = [item.text for item in tr.find_elements(By.XPATH, './/td')]
        data.append(row)
    # print(data)        

    # If not the last page, click next
    if page_num < total_pages:
        next_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/form/div[3]/div[2]/div/div/div[2]/div[2]/div/center/div/table/tfoot/tr/td/table/tbody/tr/td/div[3]/input[1]")))
        next_button.click()
        time.sleep(2)  # Allow time for the next page to load

# Create the Excel file
df = pd.DataFrame(data)
writer = pd.ExcelWriter('testdelrow.xlsx', engine='xlsxwriter')
df.to_excel(writer, sheet_name='welcome', index=False)
writer.close()

wb = openpyxl.load_workbook("testdelrow.xlsx")  # Replace with your filename
sheet = wb["welcome"]
start_row = 2
end_row = 33

if start_row < 1 or start_row > sheet.max_row or end_row > sheet.max_row:
    print("Error: Invalid row range. Please check the sheet dimensions.")
else:
    for row in range(end_row, start_row - 1, -1):  # Iterate in reverse order
        sheet.delete_rows(row)

# Save the modified workbook
wb.save("testdelrow5555.xlsx")

df = pd.read_excel("testdelrow5555.xlsx")

df[5] = df[5].apply(str)

for round_num in range(1): 
    df = df.drop(df.columns[[0,2, 3,6,7,9,11,12,13,15]],axis = 1) # , 6

print(df.columns)

df.to_excel(A+'.xlsx', index = False)

driver.quit()

print(df)
print(df[5])

print("count_row =",sheet.max_row)


